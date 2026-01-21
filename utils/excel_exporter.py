"""
Módulo para exportar datos contables a Excel
Genera reportes profesionales evaluables por profesores
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from database.queries import fetch_all


class ExcelExporter:
    """Exporta datos contables a archivos Excel"""
    
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Elimina la hoja por defecto
        self.setup_styles()
    
    def setup_styles(self):
        """Define estilos para el Excel"""
        # Colores
        self.color_header = "1F4E78"  # Azul oscuro
        self.color_subheader = "4472C4"  # Azul
        self.color_total = "D9E1F2"  # Azul claro
        self.color_error = "FF0000"  # Rojo
        self.color_success = "00B050"  # Verde
        
        # Fuentes
        self.font_header = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        self.font_subheader = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        self.font_normal = Font(name='Calibri', size=10)
        self.font_total = Font(name='Calibri', size=10, bold=True)
        
        # Rellenos
        self.fill_header = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        self.fill_subheader = PatternFill(start_color=self.color_subheader, end_color=self.color_subheader, fill_type="solid")
        self.fill_total = PatternFill(start_color=self.color_total, end_color=self.color_total, fill_type="solid")
        
        # Bordes
        self.border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def format_header(self, cell):
        """Aplica formato de encabezado a una celda"""
        cell.font = self.font_header
        cell.fill = self.fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = self.border_thin
    
    def format_subheader(self, cell):
        """Aplica formato de subencabezado"""
        cell.font = self.font_subheader
        cell.fill = self.fill_subheader
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.border_thin
    
    def format_total(self, cell, is_number=False):
        """Aplica formato de total"""
        cell.font = self.font_total
        cell.fill = self.fill_total
        cell.border = self.border_thin
        if is_number:
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = Alignment(horizontal='left')
    
    def format_number(self, cell):
        """Formatea una celda como número"""
        cell.number_format = '#,##0.00'
        cell.border = self.border_thin
        cell.font = self.font_normal
        cell.alignment = Alignment(horizontal='right')
    
    def format_text(self, cell):
        """Formatea una celda como texto"""
        cell.border = self.border_thin
        cell.font = self.font_normal
        cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    def export_plan_cuentas(self):
        """Exporta el Plan de Cuentas"""
        ws = self.wb.create_sheet("Plan de Cuentas")
        
        # Encabezado
        headers = ["Código", "Nombre", "Elemento", "Categoría", "Subcategoría", "Grupo"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos
        query = "SELECT codigo, nombre, elemento, categoria, subcategoria, grupo FROM plan_cuentas ORDER BY codigo"
        data = fetch_all(query)
        
        # Llenar datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                # Todos los campos como texto (códigos son D.CC.SS.NNNN)
                self.format_text(cell)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        
        # Fila de totales
        total_row = len(data) + 2
        ws.cell(row=total_row, column=1).value = f"Total de Cuentas: {len(data)}"
        self.format_total(ws.cell(row=total_row, column=1))
        
        return ws
    
    def export_comprobantes(self):
        """Exporta los Comprobantes y sus detalles"""
        ws = self.wb.create_sheet("Comprobantes")
        
        # Encabezado principal
        headers = ["Nº Comprobante", "Fecha", "Glosa", "Código Cuenta", "Nombre Cuenta", "Debe", "Haber"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos
        query = """
            SELECT 
                c.numero,
                c.fecha,
                c.glosa,
                d.codigo_cuenta,
                p.nombre,
                d.debe,
                d.haber
            FROM comprobantes c
            LEFT JOIN detalle_comprobantes d ON c.numero = d.numero_comprobante
            LEFT JOIN plan_cuentas p ON d.codigo_cuenta = p.codigo
            ORDER BY c.numero, d.linea
        """
        data = fetch_all(query)
        
        # Llenar datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                if col_idx == 1:  # Nº Comprobante (número)
                    cell.number_format = '0'
                    self.format_text(cell)
                elif col_idx in [6, 7]:  # Moneda
                    self.format_number(cell)
                else:  # Col 4 es código (texto D.CC.SS.NNNN)
                    self.format_text(cell)
        
        # Ajustar ancho
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Totales de Debe y Haber
        if data:
            total_row = len(data) + 2
            ws.cell(row=total_row, column=5).value = "TOTAL"
            self.format_total(ws.cell(row=total_row, column=5))
            
            # Fórmulas de suma
            debe_cell = ws.cell(row=total_row, column=6)
            haber_cell = ws.cell(row=total_row, column=7)
            debe_cell.value = f"=SUM(F2:F{total_row-1})"
            haber_cell.value = f"=SUM(G2:G{total_row-1})"
            self.format_total(debe_cell, is_number=True)
            self.format_total(haber_cell, is_number=True)
            
            # Validación: deben cuadrar
            validate_row = total_row + 1
            ws.cell(row=validate_row, column=5).value = "Cuadre (Debe=Haber)"
            ws.cell(row=validate_row, column=6).value = f"=IF(ABS(F{total_row}-G{total_row})<0.01,\"✓ CUADRA\",\"✗ NO CUADRA\")"
            
        return ws
    
    def export_libro_diario(self):
        """Exporta el Libro Diario"""
        ws = self.wb.create_sheet("Libro Diario")
        
        # Encabezado
        headers = ["Fecha", "Nº Comp.", "Código", "Cuenta", "Debe", "Haber", "Glosa"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos ordenados por fecha
        query = """
            SELECT 
                c.fecha,
                c.numero,
                d.codigo_cuenta,
                p.nombre,
                d.debe,
                d.haber,
                c.glosa
            FROM comprobantes c
            LEFT JOIN detalle_comprobantes d ON c.numero = d.numero_comprobante
            LEFT JOIN plan_cuentas p ON d.codigo_cuenta = p.codigo
            ORDER BY c.fecha, c.numero, d.linea
        """
        data = fetch_all(query)
        
        # Llenar datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                
                if col_idx == 2:  # Nº Comp (número)
                    cell.number_format = '0'
                    self.format_text(cell)
                elif col_idx in [5, 6]:  # Moneda
                    self.format_number(cell)
                else:  # Col 3 es código (texto D.CC.SS.NNNN)
                    self.format_text(cell)
        
        # Ajustar ancho
        widths = [14, 10, 12, 20, 14, 14, 30]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Totales
        if data:
            total_row = len(data) + 2
            ws.cell(row=total_row, column=4).value = "TOTALES"
            self.format_total(ws.cell(row=total_row, column=4))
            
            debe_cell = ws.cell(row=total_row, column=5)
            haber_cell = ws.cell(row=total_row, column=6)
            debe_cell.value = f"=SUM(E2:E{total_row-1})"
            haber_cell.value = f"=SUM(F2:F{total_row-1})"
            self.format_total(debe_cell, is_number=True)
            self.format_total(haber_cell, is_number=True)
        
        return ws
    
    def export_balance_comprobacion(self):
        """Exporta el Balance de Comprobación"""
        ws = self.wb.create_sheet("Balance de Comprobación")
        
        # Encabezado
        headers = ["Código", "Cuenta", "Debe", "Haber", "Saldo Deudor", "Saldo Acreedor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos con elemento para determinar naturaleza
        query = """
            SELECT 
                p.codigo,
                p.nombre,
                p.elemento,
                COALESCE(SUM(d.debe), 0) as total_debe,
                COALESCE(SUM(d.haber), 0) as total_haber
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            GROUP BY p.codigo, p.nombre, p.elemento
            HAVING (total_debe > 0 OR total_haber > 0)
            ORDER BY p.codigo
        """
        data = fetch_all(query)
        
        # Llenar datos con cálculos según naturaleza de la cuenta
        for row_idx, row_data in enumerate(data, 2):
            codigo, nombre, elemento, debe, haber = row_data
            
            ws.cell(row=row_idx, column=1).value = codigo
            ws.cell(row=row_idx, column=2).value = nombre
            ws.cell(row=row_idx, column=3).value = debe
            ws.cell(row=row_idx, column=4).value = haber
            
            # Calcular saldo según naturaleza de la cuenta
            if elemento in ['Activos', 'Gastos']:
                # Cuentas de naturaleza DEUDORA: Saldo = Debe - Haber
                saldo = debe - haber
                saldo_d = saldo if saldo > 0 else 0
                saldo_a = abs(saldo) if saldo < 0 else 0
            else:  # Pasivos, Patrimonio, Ingresos
                # Cuentas de naturaleza ACREEDORA: Saldo = Haber - Debe
                saldo = haber - debe
                saldo_a = saldo if saldo > 0 else 0
                saldo_d = abs(saldo) if saldo < 0 else 0
            
            ws.cell(row=row_idx, column=5).value = saldo_d
            ws.cell(row=row_idx, column=6).value = saldo_a
            
            # Formatear
            self.format_text(ws.cell(row=row_idx, column=1))
            self.format_text(ws.cell(row=row_idx, column=2))
            for col in [3, 4, 5, 6]:
                self.format_number(ws.cell(row=row_idx, column=col))
        
        # Totales
        if data:
            total_row = len(data) + 2
            ws.cell(row=total_row, column=1).value = "TOTALES"
            self.format_total(ws.cell(row=total_row, column=1))
            
            for col in range(3, 7):
                cell = ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{total_row-1})"
                self.format_total(cell, is_number=True)
            
            # Validación
            validate_row = total_row + 1
            ws.cell(row=validate_row, column=1).value = "Validación"
            validate_cell = ws.cell(row=validate_row, column=3)
            validate_cell.value = f"=IF(AND(ABS(C{total_row}-D{total_row})<0.01, ABS(E{total_row}-F{total_row})<0.01), \"✓ BALANCE CUADRA\", \"✗ NO CUADRA\")"
            self.format_total(validate_cell)
        
        # Ajustar ancho
        widths = [12, 25, 14, 14, 15, 15]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        return ws
    
    def export_resumen_evaluacion(self):
        """Exporta un resumen para evaluación del profesor"""
        ws = self.wb.create_sheet("Resumen Evaluación", 0)  # Primera hoja
        
        # Título
        ws.merge_cells('A1:F1')
        title = ws['A1']
        title.value = "REPORTE DE EVALUACIÓN CONTABLE"
        title.font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Información general
        row = 3
        info_labels = [
            ("Fecha de Generación:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ]
        
        for label, value in info_labels:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            self.format_text(ws.cell(row=row, column=1))
            self.format_text(ws.cell(row=row, column=2))
            row += 1
        
        # Estadísticas
        row += 1
        ws.cell(row=row, column=1).value = "ESTADÍSTICAS GENERALES"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        # Contar datos
        plan_cuentas = fetch_all("SELECT COUNT(*) FROM plan_cuentas")
        comprobantes = fetch_all("SELECT COUNT(*) FROM comprobantes")
        detalles = fetch_all("SELECT COUNT(*) FROM detalle_comprobantes")
        libro_compras = fetch_all("SELECT COUNT(*) FROM libro_compras")
        libro_ventas = fetch_all("SELECT COUNT(*) FROM libro_ventas")
        
        stats = [
            ("Cuentas en Plan de Cuentas:", plan_cuentas[0][0] if plan_cuentas else 0),
            ("Comprobantes Registrados:", comprobantes[0][0] if comprobantes else 0),
            ("Asientos Contables:", detalles[0][0] if detalles else 0),
            ("Registros en Libro de Compras:", libro_compras[0][0] if libro_compras else 0),
            ("Registros en Libro de Ventas:", libro_ventas[0][0] if libro_ventas else 0),
        ]
        
        for label, value in stats:
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value
            self.format_text(ws.cell(row=row, column=1))
            cell = ws.cell(row=row, column=2)
            cell.value = value
            cell.number_format = '0'
            self.format_text(cell)
            row += 1
        
        # Distribución por Elementos
        row += 1
        ws.cell(row=row, column=1).value = "DISTRIBUCIÓN DE CUENTAS POR ELEMENTO"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        elementos_query = "SELECT elemento, COUNT(*) FROM plan_cuentas GROUP BY elemento ORDER BY elemento"
        elementos_data = fetch_all(elementos_query)
        
        for elemento, count in elementos_data:
            ws.cell(row=row, column=1).value = f"{elemento}:"
            ws.cell(row=row, column=2).value = count
            self.format_text(ws.cell(row=row, column=1))
            cell = ws.cell(row=row, column=2)
            cell.value = count
            cell.number_format = '0'
            self.format_text(cell)
            row += 1
        
        # Validación de integridad
        row += 1
        ws.cell(row=row, column=1).value = "VALIDACIÓN DE INTEGRIDAD CONTABLE"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        debe_total = fetch_all("SELECT COALESCE(SUM(debe), 0) FROM detalle_comprobantes")
        haber_total = fetch_all("SELECT COALESCE(SUM(haber), 0) FROM detalle_comprobantes")
        
        debe = debe_total[0][0] if debe_total else 0
        haber = haber_total[0][0] if haber_total else 0
        
        ws.cell(row=row, column=1).value = "Total Débitos:"
        ws.cell(row=row, column=2).value = debe
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        ws.cell(row=row, column=1).value = "Total Créditos:"
        ws.cell(row=row, column=2).value = haber
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        ws.cell(row=row, column=1).value = "Diferencia:"
        ws.cell(row=row, column=2).value = abs(debe - haber)
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        # Estado de cuadre
        ws.cell(row=row, column=1).value = "Estado de Cuadre:"
        cuadra = "✓ CUADRADO" if abs(debe - haber) < 0.01 else "✗ NO CUADRA"
        cell = ws.cell(row=row, column=2)
        cell.value = cuadra
        cell.font = Font(name='Calibri', size=11, bold=True, 
                        color=self.color_success if "CUADRADO" in cuadra else self.color_error)
        self.format_text(ws.cell(row=row, column=1))
        row += 1
        
        # Validación ecuación contable
        row += 1
        ws.cell(row=row, column=1).value = "VALIDACIÓN ECUACIÓN CONTABLE"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        # Calcular totales por elemento
        activos_total = fetch_all("""
            SELECT COALESCE(SUM(d.debe), 0) - COALESCE(SUM(d.haber), 0)
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Activos'
        """)
        
        pasivos_total = fetch_all("""
            SELECT COALESCE(SUM(d.haber), 0) - COALESCE(SUM(d.debe), 0)
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Pasivos'
        """)
        
        patrimonio_total = fetch_all("""
            SELECT COALESCE(SUM(d.haber), 0) - COALESCE(SUM(d.debe), 0)
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Patrimonio'
        """)
        
        # Calcular utilidad del ejercicio
        from models.reportes import ReportesModel
        resultados = ReportesModel.calcular_utilidad_impuesto()
        utilidad = resultados['utilidad_ejercicio']
        
        activos = activos_total[0][0] if activos_total else 0
        pasivos = pasivos_total[0][0] if pasivos_total else 0
        patrimonio = (patrimonio_total[0][0] if patrimonio_total else 0) + utilidad
        
        ws.cell(row=row, column=1).value = "Total Activos:"
        ws.cell(row=row, column=2).value = activos
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        ws.cell(row=row, column=1).value = "Total Pasivos:"
        ws.cell(row=row, column=2).value = pasivos
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        ws.cell(row=row, column=1).value = "Total Patrimonio (incluye utilidad):"
        ws.cell(row=row, column=2).value = patrimonio
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 1
        
        ws.cell(row=row, column=1).value = "Ecuación Contable (A = P + Pat):"
        ecuacion_valida = abs(activos - (pasivos + patrimonio)) < 0.01
        ecuacion_texto = "✓ VÁLIDA" if ecuacion_valida else "✗ NO VÁLIDA"
        cell = ws.cell(row=row, column=2)
        cell.value = ecuacion_texto
        cell.font = Font(name='Calibri', size=11, bold=True, 
                        color=self.color_success if ecuacion_valida else self.color_error)
        self.format_text(ws.cell(row=row, column=1))
        row += 1
        
        # Ajustar ancho
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 25
        
        return ws
    
    def export_estado_financiero(self):
        """Exporta el Estado de Situación Financiera"""
        from config import IMPUESTO_RENTA_RATE
        
        ws = self.wb.create_sheet("Estado de Situación")
        
        # Título
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "ESTADO DE SITUACIÓN FINANCIERA"
        title.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Calcular utilidad e impuesto
        from models.reportes import ReportesModel
        resultados = ReportesModel.calcular_utilidad_impuesto()
        utilidad_ejercicio = resultados['utilidad_ejercicio']
        impuesto_por_pagar = resultados['impuesto']
        
        row = 3
        
        # Secciones del estado
        secciones = [
            ('Activo', 'Activos'),
            ('Pasivo', 'Pasivos'),
            ('Patrimonio', 'Patrimonio')
        ]
        
        totales = {}
        
        for seccion_nombre, elemento in secciones:
            # Encabezado de sección
            ws.cell(row=row, column=1).value = seccion_nombre.upper()
            self.format_subheader(ws.cell(row=row, column=1))
            row += 1
            
            # Obtener datos de la sección
            query = """
                SELECT 
                    p.codigo,
                    p.nombre,
                    COALESCE(SUM(d.debe), 0) as total_debe,
                    COALESCE(SUM(d.haber), 0) as total_haber
                FROM plan_cuentas p
                LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
                WHERE p.elemento = ?
                GROUP BY p.codigo, p.nombre
                HAVING (total_debe > 0 OR total_haber > 0)
                ORDER BY p.codigo
            """
            data = fetch_all(query, (elemento,))
            
            seccion_total = 0
            for codigo, nombre, debe, haber in data:
                # Calcular saldo según naturaleza de la cuenta
                if elemento == 'Activos':
                    # Activos: naturaleza DEUDORA (Debe - Haber)
                    saldo = debe - haber
                else:
                    # Pasivos y Patrimonio: naturaleza ACREEDORA (Haber - Debe)
                    saldo = haber - debe
                
                ws.cell(row=row, column=1).value = nombre
                ws.cell(row=row, column=2).value = saldo
                self.format_text(ws.cell(row=row, column=1))
                self.format_number(ws.cell(row=row, column=2))
                seccion_total += saldo
                row += 1
            
            # Agregar elementos calculados
            if elemento == 'Pasivos' and impuesto_por_pagar > 0:
                ws.cell(row=row, column=1).value = "Impuesto por Pagar"
                ws.cell(row=row, column=2).value = impuesto_por_pagar
                self.format_text(ws.cell(row=row, column=1))
                self.format_number(ws.cell(row=row, column=2))
                seccion_total += impuesto_por_pagar
                row += 1
            
            if elemento == 'Patrimonio' and utilidad_ejercicio != 0:
                ws.cell(row=row, column=1).value = "Utilidad del Ejercicio"
                ws.cell(row=row, column=2).value = utilidad_ejercicio
                self.format_text(ws.cell(row=row, column=1))
                self.format_number(ws.cell(row=row, column=2))
                seccion_total += utilidad_ejercicio
                row += 1
            
            # Total de sección
            total_cell = ws.cell(row=row, column=1)
            total_cell.value = f"Total {seccion_nombre}"
            self.format_total(total_cell)
            
            total_value = ws.cell(row=row, column=2)
            total_value.value = seccion_total
            self.format_total(total_value, is_number=True)
            
            totales[seccion_nombre] = seccion_total
            row += 2
        
        # Ajustar ancho
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        
        return ws
    
    def export_estado_resultados(self):
        """Exporta el Estado de Resultados según estructura NIIF"""
        from config import IMPUESTO_RENTA_RATE
        
        ws = self.wb.create_sheet("Estado de Resultados")
        
        # Título
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "ESTADO DE RESULTADOS"
        title.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        
        row = 3
        
        # INGRESOS DE ACTIVIDADES ORDINARIAS
        ws.cell(row=row, column=1).value = "INGRESOS DE ACTIVIDADES ORDINARIAS"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        query_ingresos = """
            SELECT 
                p.nombre,
                COALESCE(SUM(d.haber), 0) - COALESCE(SUM(d.debe), 0) as total_ingresos
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Ingresos'
            GROUP BY p.codigo, p.nombre
            HAVING total_ingresos <> 0
            ORDER BY p.codigo
        """
        ingresos_data = fetch_all(query_ingresos)
        total_ingresos = 0
        for nombre, monto in ingresos_data:
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = monto
            self.format_text(ws.cell(row=row, column=1))
            self.format_number(ws.cell(row=row, column=2))
            total_ingresos += monto
            row += 1
        
        # Total ingresos
        ws.cell(row=row, column=1).value = "Total Ingresos Ordinarios"
        self.format_total(ws.cell(row=row, column=1))
        total_ing = ws.cell(row=row, column=2)
        total_ing.value = total_ingresos
        self.format_total(total_ing, is_number=True)
        row += 2
        
        # COSTO DE VENTAS (usando subcategoría de Gastos)
        ws.cell(row=row, column=1).value = "COSTO DE VENTAS"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        query_costos = """
            SELECT 
                p.nombre,
                COALESCE(SUM(d.debe), 0) - COALESCE(SUM(d.haber), 0) as total_costos
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Gastos' AND p.subcategoria LIKE '%Costo%'
            GROUP BY p.codigo, p.nombre
            HAVING total_costos <> 0
            ORDER BY p.codigo
        """
        costos_data = fetch_all(query_costos)
        total_costos = 0
        for nombre, monto in costos_data:
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = monto
            self.format_text(ws.cell(row=row, column=1))
            self.format_number(ws.cell(row=row, column=2))
            total_costos += monto
            row += 1
        
        # Total costos
        ws.cell(row=row, column=1).value = "Total Costo de Ventas"
        self.format_total(ws.cell(row=row, column=1))
        total_cost = ws.cell(row=row, column=2)
        total_cost.value = total_costos
        self.format_total(total_cost, is_number=True)
        row += 2
        
        # UTILIDAD BRUTA
        utilidad_bruta = total_ingresos - total_costos
        ws.cell(row=row, column=1).value = "UTILIDAD BRUTA"
        self.format_total(ws.cell(row=row, column=1))
        ub_cell = ws.cell(row=row, column=2)
        ub_cell.value = utilidad_bruta
        self.format_total(ub_cell, is_number=True)
        row += 2
        
        # GASTOS DE ADMINISTRACIÓN
        ws.cell(row=row, column=1).value = "GASTOS DE ADMINISTRACIÓN"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        query_gastos_admin = """
            SELECT 
                p.nombre,
                COALESCE(SUM(d.debe), 0) - COALESCE(SUM(d.haber), 0) as total_gastos
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Gastos' AND p.subcategoria LIKE '%Administración%'
            GROUP BY p.codigo, p.nombre
            HAVING total_gastos <> 0
            ORDER BY p.codigo
        """
        gastos_admin_data = fetch_all(query_gastos_admin)
        total_gastos_admin = 0
        for nombre, monto in gastos_admin_data:
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = monto
            self.format_text(ws.cell(row=row, column=1))
            self.format_number(ws.cell(row=row, column=2))
            total_gastos_admin += monto
            row += 1
        
        if gastos_admin_data:
            ws.cell(row=row, column=1).value = "Total Gastos Administración"
            self.format_total(ws.cell(row=row, column=1))
            total_ga = ws.cell(row=row, column=2)
            total_ga.value = total_gastos_admin
            self.format_total(total_ga, is_number=True)
            row += 2
        
        # GASTOS DE VENTAS
        ws.cell(row=row, column=1).value = "GASTOS DE VENTAS"
        self.format_subheader(ws.cell(row=row, column=1))
        row += 1
        
        query_gastos_ventas = """
            SELECT 
                p.nombre,
                COALESCE(SUM(d.debe), 0) - COALESCE(SUM(d.haber), 0) as total_gastos
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Gastos' 
              AND p.subcategoria LIKE '%Gastos de Ventas%'
            GROUP BY p.codigo, p.nombre
            HAVING total_gastos <> 0
            ORDER BY p.codigo
        """
        gastos_ventas_data = fetch_all(query_gastos_ventas)
        total_gastos_ventas = 0
        for nombre, monto in gastos_ventas_data:
            ws.cell(row=row, column=1).value = nombre
            ws.cell(row=row, column=2).value = monto
            self.format_text(ws.cell(row=row, column=1))
            self.format_number(ws.cell(row=row, column=2))
            total_gastos_ventas += monto
            row += 1
        
        if gastos_ventas_data:
            ws.cell(row=row, column=1).value = "Total Gastos de Ventas"
            self.format_total(ws.cell(row=row, column=1))
            total_gv = ws.cell(row=row, column=2)
            total_gv.value = total_gastos_ventas
            self.format_total(total_gv, is_number=True)
            row += 2
        
        # RESULTADO OPERACIONAL
        resultado_operacional = utilidad_bruta - total_gastos_admin - total_gastos_ventas
        ws.cell(row=row, column=1).value = "RESULTADO OPERACIONAL"
        self.format_total(ws.cell(row=row, column=1))
        ro_cell = ws.cell(row=row, column=2)
        ro_cell.value = resultado_operacional
        ro_cell.font = Font(name='Calibri', size=11, bold=True)
        ro_cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        ro_cell.number_format = '#,##0.00'
        ro_cell.border = self.border_thin
        row += 2
        
        # COSTOS FINANCIEROS
        query_gastos_financieros = """
            SELECT 
                p.nombre,
                COALESCE(SUM(d.debe), 0) - COALESCE(SUM(d.haber), 0) as total_gastos
            FROM plan_cuentas p
            LEFT JOIN detalle_comprobantes d ON p.codigo = d.codigo_cuenta
            WHERE p.elemento = 'Gastos' AND p.subcategoria LIKE '%Financier%'
            GROUP BY p.codigo, p.nombre
            HAVING total_gastos <> 0
            ORDER BY p.codigo
        """
        gastos_financieros_data = fetch_all(query_gastos_financieros)
        total_gastos_financieros = 0
        
        if gastos_financieros_data:
            ws.cell(row=row, column=1).value = "COSTOS FINANCIEROS"
            self.format_subheader(ws.cell(row=row, column=1))
            row += 1
            
            for nombre, monto in gastos_financieros_data:
                ws.cell(row=row, column=1).value = nombre
                ws.cell(row=row, column=2).value = monto
                self.format_text(ws.cell(row=row, column=1))
                self.format_number(ws.cell(row=row, column=2))
                total_gastos_financieros += monto
                row += 1
            
            ws.cell(row=row, column=1).value = "Total Costos Financieros"
            self.format_total(ws.cell(row=row, column=1))
            total_gf = ws.cell(row=row, column=2)
            total_gf.value = total_gastos_financieros
            self.format_total(total_gf, is_number=True)
            row += 2
        
        # RESULTADO ANTES DE IMPUESTO
        resultado_antes_impuesto = resultado_operacional - total_gastos_financieros
        ws.cell(row=row, column=1).value = "RESULTADO ANTES DE IMPUESTO A LA RENTA"
        self.format_total(ws.cell(row=row, column=1))
        rai_cell = ws.cell(row=row, column=2)
        rai_cell.value = resultado_antes_impuesto
        rai_cell.font = Font(name='Calibri', size=11, bold=True)
        rai_cell.fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        rai_cell.number_format = '#,##0.00'
        rai_cell.border = self.border_thin
        row += 2
        
        # IMPUESTO A LA RENTA
        impuesto = resultado_antes_impuesto * IMPUESTO_RENTA_RATE if resultado_antes_impuesto > 0 else 0
        ws.cell(row=row, column=1).value = f"Gasto por Impuesto a la Renta ({int(IMPUESTO_RENTA_RATE*100)}%)"
        ws.cell(row=row, column=2).value = impuesto
        self.format_text(ws.cell(row=row, column=1))
        self.format_number(ws.cell(row=row, column=2))
        row += 2
        
        # GANANCIA (PÉRDIDA) DEL EJERCICIO
        resultado_ejercicio = resultado_antes_impuesto - impuesto
        resultado_cell = ws.cell(row=row, column=1)
        resultado_cell.value = "GANANCIA (PÉRDIDA) DEL EJERCICIO"
        resultado_cell.font = Font(name='Calibri', size=12, bold=True)
        resultado_cell.fill = PatternFill(start_color="B0D0FF", end_color="B0D0FF", fill_type="solid")
        resultado_cell.border = self.border_thin
        
        resultado_val = ws.cell(row=row, column=2)
        resultado_val.value = resultado_ejercicio
        resultado_val.number_format = '#,##0.00'
        resultado_val.font = Font(name='Calibri', size=12, bold=True)
        resultado_val.fill = PatternFill(start_color="B0D0FF", end_color="B0D0FF", fill_type="solid")
        resultado_val.border = self.border_thin
        
        # Ajustar ancho
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 18
        
        return ws
    
    def export_libro_compras(self):
        """Exporta el Libro de Compras"""
        ws = self.wb.create_sheet("Libro de Compras")
        
        # Título
        ws.merge_cells('A1:H1')
        title = ws['A1']
        title.value = "LIBRO DE COMPRAS"
        title.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezado
        headers = ["Fecha", "Proveedor", "RUT", "N° Factura", "Neto", "IVA", "Total", "N° Comp."]
        row = 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos
        query = """
            SELECT fecha, razon_social, rut_proveedor, numero_documento, 
                   neto, iva, total, numero_comprobante
            FROM libro_compras
            ORDER BY fecha, numero_documento
        """
        data = fetch_all(query)
        
        # Llenar datos
        row = 3
        for row_data in data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                
                if col_idx in [5, 6, 7]:  # Montos
                    self.format_number(cell)
                elif col_idx == 8:  # Número comprobante
                    cell.number_format = '0'
                    self.format_text(cell)
                else:
                    self.format_text(cell)
            row += 1
        
        # Totales
        if data:
            total_row = row
            ws.cell(row=total_row, column=4).value = "TOTALES"
            self.format_total(ws.cell(row=total_row, column=4))
            
            for col in [5, 6, 7]:
                cell = ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{total_row-1})"
                self.format_total(cell, is_number=True)
        
        # Ajustar ancho
        widths = [12, 25, 15, 15, 14, 14, 14, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        return ws
    
    def export_libro_ventas(self):
        """Exporta el Libro de Ventas"""
        ws = self.wb.create_sheet("Libro de Ventas")
        
        # Título
        ws.merge_cells('A1:H1')
        title = ws['A1']
        title.value = "LIBRO DE VENTAS"
        title.font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.color_header, end_color=self.color_header, fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        
        # Encabezado
        headers = ["Fecha", "Cliente", "RUT", "N° Factura", "Neto", "IVA", "Total", "N° Comp."]
        row = 2
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            self.format_header(cell)
        
        # Obtener datos
        query = """
            SELECT fecha, razon_social, rut_cliente, numero_documento, 
                   neto, iva, total, numero_comprobante
            FROM libro_ventas
            ORDER BY fecha, numero_documento
        """
        data = fetch_all(query)
        
        # Llenar datos
        row = 3
        for row_data in data:
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = value
                
                if col_idx in [5, 6, 7]:  # Montos
                    self.format_number(cell)
                elif col_idx == 8:  # Número comprobante
                    cell.number_format = '0'
                    self.format_text(cell)
                else:
                    self.format_text(cell)
            row += 1
        
        # Totales
        if data:
            total_row = row
            ws.cell(row=total_row, column=4).value = "TOTALES"
            self.format_total(ws.cell(row=total_row, column=4))
            
            for col in [5, 6, 7]:
                cell = ws.cell(row=total_row, column=col)
                cell.value = f"=SUM({get_column_letter(col)}3:{get_column_letter(col)}{total_row-1})"
                self.format_total(cell, is_number=True)
        
        # Ajustar ancho
        widths = [12, 25, 15, 15, 14, 14, 14, 12]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        return ws
    
    def save(self, filename=None):
        """Guarda el archivo Excel"""
        if filename is None:
            filename = f"Evaluacion_Contable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        self.wb.save(filename)
        return filename


def export_all_data(filename=None):
    """
    Función auxiliar para exportar todos los datos
    Retorna la ruta del archivo generado
    """
    exporter = ExcelExporter()
    # Orden de hojas optimizado para evaluación
    exporter.export_resumen_evaluacion()  # Primera hoja: resumen general
    exporter.export_plan_cuentas()
    exporter.export_comprobantes()
    exporter.export_libro_diario()
    exporter.export_libro_compras()
    exporter.export_libro_ventas()
    exporter.export_balance_comprobacion()
    exporter.export_estado_financiero()
    exporter.export_estado_resultados()
    
    return exporter.save(filename)
